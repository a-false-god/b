#!/usr/bin/env python3
"""
Catalog Importer for Prawko B.
Parses official XLSX or JSON question catalog and imports into SQLite database data/prawko.sqlite.

Features:
- Handles sheets 'katalog' (status=active) and 'W trakcie weryfikacji' (status=pending).
- Normalizes scope ('PODSTAWOWY', 'SPECJALISTYCZNY', fixing typos like 'Specajlistyczny').
- Determines type: 'TN' if correct in {'T', 'N'}, else 'ABC'.
- Categorizes correctly as JSON array (e.g. ["A", "B"]).
- Re-import performs a diff by question 'id' (insert new, update changed), preserving question_classification and answer_events.
"""

import argparse
import json
import sqlite3
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None


def normalize_scope(raw_scope: str) -> str:
    if not raw_scope:
        return "PODSTAWOWY"
    val = str(raw_scope).strip().upper()
    if "SPECJAL" in val or "SPECAJ" in val:
        return "SPECJALISTYCZNY"
    if "PODST" in val:
        return "PODSTAWOWY"
    return val


def determine_media_kind(media_name: str | None) -> str:
    if not media_name or not str(media_name).strip():
        return "none"
    ext = Path(str(media_name).strip()).suffix.lower()
    if ext in [".wmv", ".mp4", ".mov", ".avi"]:
        return "video"
    if ext in [".jpg", ".jpeg", ".webp", ".png"]:
        return "image"
    return "none"


def normalize_categories(raw_cats: str) -> str:
    if not raw_cats:
        return json.dumps([])
    cats = [c.strip().upper() for c in str(raw_cats).split(",") if c.strip()]
    return json.dumps(cats)


def parse_xlsx_catalog(xlsx_path: Path) -> list[dict]:
    if openpyxl is None:
        raise ImportError("openpyxl package is required to parse XLSX catalog files.")

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
    questions = []

    sheet_status_map = {
        "katalog": "active",
        "W trakcie weryfikacji": "pending"
    }

    for sheet_name, status in sheet_status_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            continue

        header = rows[0]
        for row in rows[1:]:
            if not row or row[1] is None:
                continue

            try:
                q_id = int(row[1])
            except (ValueError, TypeError):
                continue

            lp = int(row[0]) if row[0] is not None and str(row[0]).isdigit() else None
            q_pl = str(row[2]).strip() if row[2] else ""
            a_pl = str(row[3]).strip() if len(row) > 3 and row[3] else None
            b_pl = str(row[4]).strip() if len(row) > 4 and row[4] else None
            c_pl = str(row[5]).strip() if len(row) > 5 and row[5] else None

            correct = str(row[6]).strip().upper() if len(row) > 6 and row[6] else ""
            media = str(row[7]).strip() if len(row) > 7 and row[7] else None
            media_kind = determine_media_kind(media)

            scope = normalize_scope(row[8] if len(row) > 8 else "")
            points = int(row[9]) if len(row) > 9 and row[9] is not None and str(row[9]).isdigit() else 1
            categories_str = normalize_categories(row[10] if len(row) > 10 else "")

            q_type = "TN" if correct in ("T", "N") else "ABC"

            # Optional translations
            pjm_q = str(row[11]).strip() if len(row) > 11 and row[11] else None
            q_en = str(row[15]).strip() if len(row) > 15 and row[15] else None
            a_en = str(row[16]).strip() if len(row) > 16 and row[16] else None
            b_en = str(row[17]).strip() if len(row) > 17 and row[17] else None
            c_en = str(row[18]).strip() if len(row) > 18 and row[18] else None

            q_de = str(row[19]).strip() if len(row) > 19 and row[19] else None
            a_de = str(row[20]).strip() if len(row) > 20 and row[20] else None
            b_de = str(row[21]).strip() if len(row) > 21 and row[21] else None
            c_de = str(row[22]).strip() if len(row) > 22 and row[22] else None

            q_ua = str(row[23]).strip() if len(row) > 23 and row[23] else None
            a_ua = str(row[24]).strip() if len(row) > 24 and row[24] else None
            b_ua = str(row[25]).strip() if len(row) > 25 and row[25] else None
            c_ua = str(row[26]).strip() if len(row) > 26 and row[26] else None

            questions.append({
                "id": q_id,
                "lp": lp,
                "scope": scope,
                "points": points,
                "type": q_type,
                "correct": correct,
                "media": media,
                "media_kind": media_kind,
                "categories": categories_str,
                "status": status,
                "q_pl": q_pl, "a_pl": a_pl, "b_pl": b_pl, "c_pl": c_pl,
                "q_en": q_en, "a_en": a_en, "b_en": b_en, "c_en": c_en,
                "q_de": q_de, "a_de": a_de, "b_de": b_de, "c_de": c_de,
                "q_ua": q_ua, "a_ua": a_ua, "b_ua": b_ua, "c_ua": c_ua,
                "pjm_q": pjm_q
            })

    wb.close()
    return questions


def import_questions_to_db(questions: list[dict], db_path: Path) -> tuple[int, int]:
    """Upsert questions into SQLite database."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Ensure schema exists
    migration_sql = Path(__file__).resolve().parent / "migrate_001.sql"
    if migration_sql.exists():
        cursor.executescript(migration_sql.read_text(encoding="utf-8"))

    inserted = 0
    updated = 0

    for q in questions:
        cursor.execute("SELECT id FROM questions WHERE id = ?", (q["id"],))
        exists = cursor.fetchone()

        if exists:
            updated += 1
            cursor.execute(
                """
                UPDATE questions SET
                  lp=?, scope=?, points=?, type=?, correct=?, media=?, media_kind=?,
                  categories=?, status=?, q_pl=?, a_pl=?, b_pl=?, c_pl=?,
                  q_en=?, a_en=?, b_en=?, c_en=?, q_de=?, a_de=?, b_de=?, c_de=?,
                  q_ua=?, a_ua=?, b_ua=?, c_ua=?, pjm_q=?
                WHERE id=?
                """,
                (
                    q["lp"], q["scope"], q["points"], q["type"], q["correct"], q["media"], q["media_kind"],
                    q["categories"], q["status"], q["q_pl"], q["a_pl"], q["b_pl"], q["c_pl"],
                    q["q_en"], q["a_en"], q["b_en"], q["c_en"], q["q_de"], q["a_de"], q["b_de"], q["c_de"],
                    q["q_ua"], q["a_ua"], q["b_ua"], q["c_ua"], q["pjm_q"],
                    q["id"]
                )
            )
        else:
            inserted += 1
            cursor.execute(
                """
                INSERT INTO questions (
                  id, lp, scope, points, type, correct, media, media_kind,
                  categories, status, q_pl, a_pl, b_pl, c_pl,
                  q_en, a_en, b_en, c_en, q_de, a_de, b_de, c_de,
                  q_ua, a_ua, b_ua, c_ua, pjm_q
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    q["id"], q["lp"], q["scope"], q["points"], q["type"], q["correct"], q["media"], q["media_kind"],
                    q["categories"], q["status"], q["q_pl"], q["a_pl"], q["b_pl"], q["c_pl"],
                    q["q_en"], q["a_en"], q["b_en"], q["c_en"], q["q_de"], q["a_de"], q["b_de"], q["c_de"],
                    q["q_ua"], q["a_ua"], q["b_ua"], q["c_ua"], q["pjm_q"]
                )
            )

    conn.commit()
    conn.close()
    return inserted, updated


def main():
    parser = argparse.ArgumentParser(description="Import question catalog into SQLite")
    parser.add_argument("--xlsx", help="Path to catalog Excel file")
    parser.add_argument("--json", help="Path to catalog JSON file")
    parser.add_argument("--db", default="data/prawko.sqlite", help="Target SQLite file")
    args = parser.parse_args()

    project_root = Path(__file__).resolve().parent.parent
    db_path = Path(args.db) if os.path.isabs(args.db) else project_root / args.db
    db_path.parent.mkdir(parents=True, exist_ok=True)

    questions = []
    if args.xlsx:
        xlsx_path = Path(args.xlsx) if os.path.isabs(args.xlsx) else project_root / args.xlsx
        questions = parse_xlsx_catalog(xlsx_path)
    elif args.json:
        json_path = Path(args.json) if os.path.isabs(args.json) else project_root / args.json
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            questions = data if isinstance(data, list) else data.get("questions", [])

    if not questions:
        print("No questions provided or parsed.")
        sys.exit(1)

    ins, upd = import_questions_to_db(questions, db_path)
    print(f"Catalog import completed: {ins} inserted, {upd} updated into {db_path}")


if __name__ == "__main__":
    import os
    main()
